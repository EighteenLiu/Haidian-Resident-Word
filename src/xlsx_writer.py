from __future__ import annotations

import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

from .models import CaseRecord, ReportStats, RoadMatchResult, VillageEntry
from .utils import display_village_name, normalize_header, normalize_text


DETAIL_HEADERS = [
    "编号",
    "案件来源",
    "案件状态",
    "案件分类",
    "上报单位",
    "上报时间",
    "截止时间",
    "结案时间",
    "小区",
    "案件描述",
    "案件标签",
    "道路",
    "公园",
    "检查点位\n（1级）",
    "检查点位\n（2级）",
    "检查点位\n（3级）",
    "检查指标\n（1级）",
    "检查指标\n（2级）",
    "检查指标\n（3级）",
    "街镇分中心",
    "区委办局",
    "区委办局二级单位",
    "作业单位",
    "作业单位二级",
    "整改责任单位",
    "处置部门名称\n（一级部门）",
    "整改时间\n（二级部门）",
    "是否按期整改",
    "经纬度",
    "商户编号",
    "商户名称",
    "上报扣分案件",
    "解决案件库",
    "按期整改",
    "超期整改",
    "超期未整改(已整改未审核)",
    "超期未整改",
]


def _record_value(record: CaseRecord, header: str):
    mapping = {
        "编号": record.case_id,
        "案件来源": record.source,
        "案件状态": record.status,
        "案件分类": record.case_category,
        "上报单位": record.report_unit,
        "上报时间": record.report_time,
        "截止时间": record.deadline_time,
        "结案时间": record.close_time,
        "小区": record.community,
        "案件描述": record.description,
        "案件标签": record.tags,
        "道路": record.road,
        "公园": record.park,
        "检查点位(1级)": record.point_level_1,
        "检查点位(2级)": record.point_level_2,
        "检查点位(3级)": record.point_level_3,
        "检查指标(1级)": record.indicator_level_1,
        "检查指标(2级)": record.indicator_level_2,
        "检查指标(3级)": record.indicator_level_3,
        "街镇分中心": record.street_center,
        "区委办局": record.district_department,
        "区委办局二级单位": record.district_department_level_2,
        "作业单位": record.operation_unit,
        "作业单位二级": record.operation_unit_level_2,
        "整改责任单位": record.rectification_unit,
        "处置部门名称(一级部门)": record.disposal_department,
        "整改时间(二级部门)": record.rectification_time,
        "是否按期整改": record.on_time_rectification,
        "经纬度": record.coordinate,
        "商户编号": record.merchant_id,
        "商户名称": record.merchant_name,
        "上报扣分案件": record.is_deduct_case,
        "解决案件库": record.is_solved,
        "按期整改": record.raw.get("按期整改", ""),
        "超期整改": record.raw.get("超期整改", ""),
        "超期未整改(已整改未审核)": record.raw.get("超期未整改(已整改未审核)", ""),
        "超期未整改": record.raw.get("超期未整改", ""),
    }
    return mapping.get(normalize_header(header), "")


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    if source_row == target_row:
        return
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _set_value(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left = ws.cell(merged_range.min_row, merged_range.min_col)
                if top_left.row == row and top_left.column == col:
                    top_left.value = value
                return
        return
    cell.value = value


def write_statistics_xlsx(
    template_path: Path | None,
    output_path: Path,
    stats: ReportStats,
    records: list[CaseRecord],
    villages: list[VillageEntry],
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if template_path and Path(template_path).exists():
        wb = load_workbook(template_path)
    else:
        wb = Workbook()
        wb.active.title = "sheet1"
    ws = wb["sheet1"] if "sheet1" in wb.sheetnames else wb.active

    _write_sheet1(ws, stats, villages)
    _write_detail_sheet(wb, records)
    _write_unmapped_sheet(wb, stats)
    wb.save(output_path)
    return output_path


def _write_sheet1(ws, stats: ReportStats, villages: list[VillageEntry]) -> None:
    if ws.max_row < 6:
        ws.append(["海淀区农村人居环境检查情况统计表"])
        ws.append(["序号", "乡镇", "村"])
        ws.append([])
        ws.append([])
        ws.append([])
        ws.append(["汇总"])
    while ws.max_row < 6 + len(villages):
        ws.append([])

    # Disable old scoring formulas/results in the generated workbook.
    for row in range(6, max(ws.max_row, 6 + len(villages)) + 1):
        for col in range(4, 9):
            _set_value(ws, row, col, None)

    village_row_by_key = {}
    for idx, entry in enumerate(villages, start=7):
        if idx > ws.max_row:
            ws.append([])
        if idx != 7:
            _copy_row_style(ws, 7, idx, max(ws.max_column, 29))
        _set_value(ws, idx, 1, idx - 6)
        _set_value(ws, idx, 2, entry.town_name)
        _set_value(ws, idx, 3, display_village_name(entry.village_name))
        village_row_by_key[(normalize_text(entry.town_name), normalize_text(entry.village_name))] = idx

    subcategory_by_col = {}
    for col in range(10, min(ws.max_column, 28) + 1):
        header = ws.cell(3, col).value or ws.cell(4, col).value
        subcategory_by_col[col] = normalize_text(header)
        _set_value(ws, 6, col, f"=SUM({ws.cell(7, col).coordinate}:{ws.cell(6 + len(villages), col).coordinate})")

    for entry in villages:
        row_idx = village_row_by_key[(normalize_text(entry.town_name), normalize_text(entry.village_name))]
        counts = stats.village_indicator_counts.get((entry.town_name, entry.village_name), {})
        normalized_counts = {normalize_text(k): v for k, v in counts.items()}
        for col, subcategory in subcategory_by_col.items():
            _set_value(ws, row_idx, col, normalized_counts.get(subcategory, 0))


def _write_detail_sheet(wb, records: list[CaseRecord]) -> None:
    if "数据明细" in wb.sheetnames:
        ws = wb["数据明细"]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        ws = wb.create_sheet("数据明细")
    for col, header in enumerate(DETAIL_HEADERS, 1):
        ws.cell(1, col).value = header
    for row_idx, record in enumerate(records, 2):
        if row_idx > 2:
            _copy_row_style(ws, 2, row_idx, len(DETAIL_HEADERS))
        for col, header in enumerate(DETAIL_HEADERS, 1):
            ws.cell(row_idx, col).value = _record_value(record, header)


def _write_unmapped_sheet(wb, stats: ReportStats) -> None:
    if "未映射指标" in wb.sheetnames:
        ws = wb["未映射指标"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("未映射指标")
    ws.append(["指标小类", "数量"])
    for name, count in sorted(stats.unmapped_indicator_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([name, count])


def write_road_checklist(path: Path, results: list[RoadMatchResult]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "道路匹配校验"
    ws.append(["状态", "编号", "街镇分中心", "检查点位（2级）", "检查点位（3级）", "道路", "命中道路", "说明"])
    for result in results:
        if result.status == "MATCHED":
            continue
        entry = result.matched_entry
        ws.append([
            result.status,
            result.record.case_id,
            result.record.street_center,
            result.record.point_level_2,
            result.record.point_level_3,
            result.record.road,
            entry.road if entry else "",
            result.message,
        ])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
