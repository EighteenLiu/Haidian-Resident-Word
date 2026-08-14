from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import CaseRecord, IndicatorMapping, RoadLedgerEntry, VillageEntry
from .utils import normalize_header, normalize_text, parse_datetime


def _header_map(ws, row_idx: int = 1) -> dict[str, int]:
    return {normalize_header(ws.cell(row_idx, col).value): col for col in range(1, ws.max_column + 1) if normalize_header(ws.cell(row_idx, col).value)}


def _find_header_row(ws, required: set[str], max_rows: int = 8) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        headers = _header_map(ws, row_idx)
        if required.issubset(headers):
            return row_idx, headers
    raise ValueError(f"工作表 {ws.title} 缺少必需字段：{', '.join(sorted(required))}")


def _value(row, headers: dict[str, int], name: str, default: Any = "") -> Any:
    col = headers.get(normalize_header(name))
    if not col:
        return default
    return row[col - 1]


def read_main_records(path: Path, sheet_name: str | None = None) -> list[CaseRecord]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb["数据"] if "数据" in wb.sheetnames else wb.active
    required = {"月份", "编号", "案件状态", "上报时间", "检查点位(1级)", "检查点位(2级)", "检查点位(3级)", "检查指标(3级)", "街镇分中心", "上报扣分案件", "解决案件库"}
    header_row, headers = _find_header_row(ws, required)
    records: list[CaseRecord] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            continue
        raw = {name: row[col - 1] for name, col in headers.items() if col - 1 < len(row)}
        record = CaseRecord(
            month=str(_value(row, headers, "月份", "") or "").strip(),
            remark=str(_value(row, headers, "备注", "") or "").strip(),
            case_id=str(_value(row, headers, "编号", "") or "").strip(),
            source=str(_value(row, headers, "案件来源", "") or "").strip(),
            status=str(_value(row, headers, "案件状态", "") or "").strip(),
            case_category=str(_value(row, headers, "案件分类", "") or "").strip(),
            report_unit=str(_value(row, headers, "上报单位", "") or "").strip(),
            report_time=parse_datetime(_value(row, headers, "上报时间")),
            deadline_time=parse_datetime(_value(row, headers, "截止时间")),
            close_time=parse_datetime(_value(row, headers, "结案时间")),
            area=str(_value(row, headers, "区域", "") or "").strip(),
            community=str(_value(row, headers, "小区", "") or "").strip(),
            description=str(_value(row, headers, "案件描述", "") or "").strip(),
            tags=str(_value(row, headers, "案件标签", "") or "").strip(),
            road=str(_value(row, headers, "道路", "") or "").strip(),
            park=str(_value(row, headers, "公园", "") or "").strip(),
            point_level_1=str(_value(row, headers, "检查点位(1级)", "") or "").strip(),
            point_level_2=str(_value(row, headers, "检查点位(2级)", "") or "").strip(),
            point_level_3=str(_value(row, headers, "检查点位(3级)", "") or "").strip(),
            indicator_level_1=str(_value(row, headers, "检查指标(1级)", "") or "").strip(),
            indicator_level_2=str(_value(row, headers, "检查指标(2级)", "") or "").strip(),
            indicator_level_3=str(_value(row, headers, "检查指标(3级)", "") or "").strip(),
            street_center=str(_value(row, headers, "街镇分中心", "") or "").strip(),
            district_department=str(_value(row, headers, "区委办局", "") or "").strip(),
            district_department_level_2=str(_value(row, headers, "区委办局二级单位", "") or "").strip(),
            operation_unit=str(_value(row, headers, "作业单位", "") or "").strip(),
            operation_unit_level_2=str(_value(row, headers, "作业单位二级", "") or "").strip(),
            rectification_unit=str(_value(row, headers, "整改责任单位", "") or "").strip(),
            disposal_department=str(_value(row, headers, "处置部门名称(一级部门)", "") or "").strip(),
            rectification_time=parse_datetime(_value(row, headers, "整改时间(二级部门)")),
            on_time_rectification=str(_value(row, headers, "是否按期整改", "") or "").strip(),
            coordinate=str(_value(row, headers, "经纬度", "") or "").strip(),
            merchant_id=str(_value(row, headers, "商户编号", "") or "").strip(),
            merchant_name=str(_value(row, headers, "商户名称", "") or "").strip(),
            is_deduct_case=str(_value(row, headers, "上报扣分案件", "") or "").strip(),
            is_solved=str(_value(row, headers, "解决案件库", "") or "").strip(),
            raw=raw,
        )
        records.append(record)
    return records


def read_dictionary(path: Path) -> tuple[list[VillageEntry], dict[str, IndicatorMapping]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    village_ws = wb["人居村"]
    indicator_ws = wb["人居指标"]
    _, village_headers = _find_header_row(village_ws, {"街镇名称", "人居村"})
    villages: list[VillageEntry] = []
    for row in village_ws.iter_rows(min_row=2, values_only=True):
        town = str(_value(row, village_headers, "街镇名称", "") or "").strip()
        village = str(_value(row, village_headers, "人居村", "") or "").strip()
        if town and village:
            villages.append(VillageEntry(town, village))

    _, indicator_headers = _find_header_row(indicator_ws, {"大类名称", "次大类", "小类名称"})
    mappings: dict[str, IndicatorMapping] = {}
    for row in indicator_ws.iter_rows(min_row=2, values_only=True):
        category = str(_value(row, indicator_headers, "大类名称", "") or "").strip()
        subcategory = str(_value(row, indicator_headers, "次大类", "") or "").strip()
        indicator = str(_value(row, indicator_headers, "小类名称", "") or "").strip()
        if category and indicator:
            mappings[normalize_text(indicator)] = IndicatorMapping(category, subcategory, indicator)
    return villages, mappings


def _road_keys(street: str, village: str, road: str) -> set[str]:
    keys = set()
    ns, nv, nr = normalize_text(street), normalize_text(village), normalize_text(road)
    if ns and nv and nr:
        keys.add(ns + "|" + nv + "|" + nr)
    if nv and nr:
        keys.add(nv + "|" + nr)
    return keys


def read_road_ledger(path: Path, sheet_name: str | None = None) -> list[RoadLedgerEntry]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    header_row, headers = _find_header_row(ws, {"所属街道", "小区(村)名称", "道路", "备注-别名"})
    entries: list[RoadLedgerEntry] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        street = str(_value(row, headers, "所属街道", "") or "").strip()
        village = str(_value(row, headers, "小区(村)名称", "") or "").strip()
        road = str(_value(row, headers, "道路", "") or "").strip()
        alias = str(_value(row, headers, "备注-别名", "") or "").strip()
        if not (street and village and road):
            continue
        keys = _road_keys(street, village, road)
        for item in re.split(r"[;；、,，/]+", alias):
            item = item.strip()
            if item:
                keys.update(_road_keys(street, village, item))
        entries.append(RoadLedgerEntry(street, village, road, alias, keys))
    return entries


DETAIL_EXCLUDED_HEADERS = {"月份", "备注", "序列", "整改责任单位-报告"}


def source_detail_headers(path: Path, sheet_name: str | None = None) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb["数据"] if "数据" in wb.sheetnames else wb.active
    header_row, headers = _find_header_row(ws, {"月份", "编号", "检查点位(1级)", "检查指标(3级)"})
    ordered = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value is None:
            continue
        if normalize_header(value) not in {normalize_header(x) for x in DETAIL_EXCLUDED_HEADERS}:
            ordered.append(str(value))
    return ordered
