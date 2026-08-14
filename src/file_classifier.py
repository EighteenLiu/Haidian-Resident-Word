from __future__ import annotations

from dataclasses import dataclass
from enum import Enum
from pathlib import Path

from openpyxl import load_workbook

from .utils import normalize_header


class FileType(str, Enum):
    MAIN_DATA = "主数据"
    DICTIONARY = "村指标字典"
    ROAD_LEDGER = "道路台账"
    UNKNOWN = "未知文件"


@dataclass
class ClassifiedFile:
    path: Path
    file_type: FileType
    sheet_name: str | None = None
    reason: str = ""


MAIN_HEADERS = {
    "月份",
    "编号",
    "案件来源",
    "案件状态",
    "案件分类",
    "上报时间",
    "检查点位(1级)",
    "检查点位(2级)",
    "检查点位(3级)",
    "检查指标(1级)",
    "检查指标(2级)",
    "检查指标(3级)",
    "街镇分中心",
    "上报扣分案件",
    "解决案件库",
}

ROAD_HEADERS = {"所属街道", "小区(村)名称", "道路", "备注-别名"}


def _sheet_headers(ws, max_rows: int = 5) -> tuple[int, set[str]]:
    best_row = 1
    best_headers: set[str] = set()
    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        headers = {normalize_header(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)}
        headers.discard("")
        if len(headers) > len(best_headers):
            best_row = row_idx
            best_headers = headers
    return best_row, best_headers


def _has_headers(ws, required: set[str], max_rows: int = 5) -> bool:
    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        headers = {normalize_header(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)}
        if required.issubset(headers):
            return True
    return False


def classify_excel_file(path: Path) -> ClassifiedFile:
    path = Path(path)
    if path.name.startswith("~$"):
        return ClassifiedFile(path, FileType.UNKNOWN, reason="Excel 临时锁文件")
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return ClassifiedFile(path, FileType.UNKNOWN, reason=f"无法打开：{exc}")

    sheet_names = set(wb.sheetnames)
    if {"人居村", "人居指标"}.issubset(sheet_names):
        village_headers = _sheet_headers(wb["人居村"])[1]
        indicator_headers = _sheet_headers(wb["人居指标"])[1]
        if {"街镇名称", "人居村"}.issubset(village_headers) and {"大类名称", "次大类", "小类名称"}.issubset(indicator_headers):
            return ClassifiedFile(path, FileType.DICTIONARY, reason="包含人居村和人居指标工作表")

    for ws in wb.worksheets:
        _, headers = _sheet_headers(ws)
        if len(MAIN_HEADERS & headers) >= 13 and ws.max_column >= 35:
            return ClassifiedFile(path, FileType.MAIN_DATA, ws.title, "包含现场检查主数据表头")
        if _has_headers(ws, ROAD_HEADERS):
            return ClassifiedFile(path, FileType.ROAD_LEDGER, ws.title, "包含道路台账表头")

    return ClassifiedFile(path, FileType.UNKNOWN, reason="未匹配到已知表头结构")


def classify_files(paths: list[Path]) -> list[ClassifiedFile]:
    return [classify_excel_file(Path(p)) for p in paths if not Path(p).name.startswith("~$")]


def require_single(classified: list[ClassifiedFile], file_type: FileType) -> ClassifiedFile:
    hits = [item for item in classified if item.file_type == file_type]
    if not hits:
        raise ValueError(f"缺少{file_type.value}文件")
    if len(hits) > 1:
        names = "、".join(item.path.name for item in hits)
        raise ValueError(f"{file_type.value}文件识别到多个：{names}")
    return hits[0]
